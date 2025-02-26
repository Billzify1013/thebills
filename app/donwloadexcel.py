import openpyxl
from .models import *
import calendar
from io import BytesIO
import xlwt
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import EventBookGuest  # Import your EventBookGuest model here

def exceldatapage(request):
    try:
        if request.user.is_authenticated:
            # Validate the month
            user=request.user
            
            return render(request,'showexceldata.html',{'active_page': 'exceldatapages ',})
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)    
    
         
@login_required
def generate_invoice_excel(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            month = int(request.POST.get('monthnumber'))
            if not 1 <= month <= 12:
                messages.error(request, "Select Correct month")
                return redirect('exceldatapage')

            # Filter invoices based on the given month and invoice_status
            invoices = Invoice.objects.filter(
                invoice_date__month=month,
                invoice_status=True,
                vendor=user
            ).order_by('invoice_number')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet(f"{calendar.month_name[month]} Invoices")

            # Define the header
            headers = ["Invoice Number", "Guest Name", "Phone", "Source" ,"Check-in Date", "Check-out Date",
                    "Tax Amount", "Grand Total Amount",  "Tax Type", "GST Number"]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]

            # Write the invoice data
            for row_idx, invoice in enumerate(invoices, start=1):
                guest = invoice.customer
                row = [
                    str(invoice.invoice_number),  # Convert invoice_number to string to handle longer values
                    guest.guestname,
                    str(guest.guestphome),  # Convert guestphome to string
                    guest.channel,
                    guest.checkindate.strftime('%Y-%m-%d %H:%M'),  # Format with date and time
                    guest.checkoutdate.strftime('%Y-%m-%d %H:%M'),  # Format with date and time
                    float(invoice.gst_amount * 2),  # Example calculation, adjust as needed
                    float(invoice.grand_total_amount),
                    invoice.taxtype,
                    invoice.customer_gst_number,
                ]

                for col_idx, cell_value in enumerate(row):
                    sheet.write(row_idx, col_idx, cell_value)

                    # Update max_col_widths for the current column if necessary
                    max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 8000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{calendar.month_name[month]}_invoices.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)    
    
         
import calendar
import xlwt


@login_required
def generate_eventinvoice_excel(request):
    try:
        if request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            month = int(request.POST.get('monthnumber'))
            if not 1 <= month <= 12:
                messages.error(request, "Please select a valid month.")
                return redirect('exceldatapage')

            # Filter EventBookGuest entries for the given month and user (vendor)
            invoices = EventBookGuest.objects.filter(
                vendor=user,
                invoice_date__month=month,
                status=True
            ).order_by('invoice_number')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet(f"{calendar.month_name[month]} Invoices")

            # Define the header
            headers = ["Invoice Number", "Start Date", "End Date", "Event Name", "Customer Name", "Email", 
                    "Contact", "Address", "GST Number",  "Tax Amount", "Grand Total Amount", "Tax Type"]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]

            # Write the invoice data
            for row_idx, invoice in enumerate(invoices, start=1):
                row = [
                    str(invoice.invoice_number),  # Convert invoice_number to string to handle longer values
                    invoice.start_date.strftime('%Y-%m-%d') if invoice.start_date else '',
                    invoice.end_date.strftime('%Y-%m-%d') if invoice.end_date else '',
                    invoice.event.eventname,  # Assuming you have an Events model linked
                    invoice.customername,
                    invoice.guestemail,
                    str(invoice.customer_contact),  # Convert customer_contact to string
                    invoice.customeraddress,
                    invoice.customergst,
                    float(invoice.taxamount),
                    float(invoice.Grand_total_amount),
                    invoice.taxtype,
                ]

                for col_idx, cell_value in enumerate(row):
                    sheet.write(row_idx, col_idx, cell_value)

                    # Update max_col_widths for the current column if necessary
                    max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 8000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{calendar.month_name[month]}Events_invoices.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)    
    


def generate_aminitiesinvoice_excel(request):
    try:
        if request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            month = int(request.POST.get('monthnumber'))
            if not 1 <= month <= 12:
                messages.error(request, "Please select a valid month.")
                return redirect('exceldatapage')

            # Filter EventBookGuest entries for the given month and user (vendor)
            invoices = AminitiesInvoice.objects.filter(
                vendor=user,
                invoicedate__month=month,
                sattle=True
            ).order_by('invoicenumber')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet(f"{calendar.month_name[month]} Invoices")

            # Define the header
            headers = ["Invoice Number", "Invoice Date", "Customer name", "Customer Phone",  "Customer Email", 
                     "Customer Address", "Customer Gst NO", "Customer Company ", "Total Amount", "Discount Amount" , "SubTotal Amount" , "Tax Amount", "Grand Total Amount", "Tax Type"]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]

            # Write the invoice data
            for row_idx, invoice in enumerate(invoices, start=1):
                row = [
                    str(invoice.invoicenumber),  # Convert invoice_number to string to handle longer values
                    invoice.invoicedate.strftime('%Y-%m-%d') if invoice.invoicedate else '',
                    invoice.customername,  
                    str(invoice.customercontact),
                    invoice.customeremail,
                    invoice.customeraddress,  # Convert customer_contact to string
                    invoice.customergst,
                    invoice.customercompany,
                    float(invoice.total_item_amount),
                    float(invoice.discount_amount),
                    float(invoice.subtotal_amount),
                    float(invoice.gst_amount*2),
                    float(invoice.grand_total_amount),
                    invoice.taxtype,
                ]

                for col_idx, cell_value in enumerate(row):
                    sheet.write(row_idx, col_idx, cell_value)

                    # Update max_col_widths for the current column if necessary
                    max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 8000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{calendar.month_name[month]}Aminitines_invoices.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)    
    




def generate_purchesinvoice_excel(request):
    try:
        if request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            month = int(request.POST.get('monthnumber'))
            if not 1 <= month <= 12:
                messages.error(request, "Please select a valid month.")
                return redirect('exceldatapage')

            # Filter EventBookGuest entries for the given month and user (vendor)
            invoices = Supplier.objects.filter(
                vendor=user,
                invoicedate__month=month,
                sattle=True
            ).order_by('invoicenumber')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet(f"{calendar.month_name[month]} Purches Invoices")

            # Define the header
            headers = ["Purches Invoice Number", "Purches Invoice Date", "Supplier name", "Supplier Phone",  "Supplier Email", 
                     "Supplier Address", "Supplier Gst NO", "Supplier Company ", "Total Amount", "Discount Amount" , "SubTotal Amount" , "Tax Amount", "Grand Total Amount", "Tax Type"]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]

            # Write the invoice data
            for row_idx, invoice in enumerate(invoices, start=1):
                row = [
                    str(invoice.invoicenumber),  # Convert invoice_number to string to handle longer values
                    invoice.invoicedate.strftime('%Y-%m-%d') if invoice.invoicedate else '',
                    invoice.customername,  
                    str(invoice.customercontact),
                    invoice.customeremail,
                    invoice.customeraddress,  # Convert customer_contact to string
                    invoice.customergst,
                    invoice.companyname,
                    float(invoice.total_item_amount),
                    float(invoice.discount_amount),
                    float(invoice.subtotal_amount),
                    float(invoice.gst_amount*2),
                    float(invoice.grand_total_amount),
                    invoice.taxtype,
                ]

                for col_idx, cell_value in enumerate(row):
                    sheet.write(row_idx, col_idx, cell_value)

                    # Update max_col_widths for the current column if necessary
                    max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 8000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{calendar.month_name[month]}Purches_invoices.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)





def generate_form_purchesinvoice_excel(request):
    try:
        if request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')

            # Filter EventBookGuest entries for the given month and user (vendor)
            invoices = Supplier.objects.filter(
                vendor=user,
                invoicedate__range=[startdate, enddate],
                sattle=True,
                unpaid=False,
            ).order_by('invoicenumber')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet("Purches Invoices")

            # Define the header
            headers = [
                "Supplier State", "Supplier GSTIN", "Invoice Number", "Invoice Date", "Supplier Name",
                "Taxable Value", "HSN Code", "CGST %","CGST Amount", "SGST %", "SGST Amount", "IGST %",  "IGST Amount"
            ]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]
            row_idx = 1  # Start writing data from the second row

            # Write the invoice data
            for invoice in invoices:
                # Get the HSN data for the invoice
                hsndatas = taxSlabpurchase.objects.filter(vendor=user, invoice=invoice)

                if hsndatas.exists():
                    first_entry = True  # Flag to check if this is the first entry for the invoice
                    taxtype = invoice.taxtype  # Variable to hold the taxtype (GST or IGST)

                    

                    # Loop through the HSN data for the invoice
                    for hsndata in hsndatas:
                        # If it's the first entry, add all invoice details
                        if first_entry:
                            row = [
                                invoice.state, invoice.customergst, str(invoice.invoicenumber),
                                invoice.invoicedate.strftime('%Y-%m-%d') if invoice.invoicedate else '',
                                invoice.customername, hsndata.taxableamount, hsndata.tax_hsnsac_name,
                                hsndata.cgst if taxtype == 'GST' else '',  # For GST, include CGST; else leave empty
                                hsndata.cgst_amount if taxtype == 'GST' else '',  # For GST, include CGST amount; else leave empty
                                hsndata.sgst if taxtype == 'GST' else '',  # For GST, include SGST; else leave empty
                                hsndata.sgst_amount if taxtype == 'GST' else '',  # For GST, include SGST amount; else leave empty
                                hsndata.cgst*2 if taxtype == 'IGST' else '',  # For IGST, include IGST; else leave empty
                               
                                hsndata.cgst_amount*2 if taxtype == 'IGST' else '',  # For IGST, include IGST amount; else leave empty
                            ]
                            first_entry = False  # After the first entry, set the flag to false
                        else:
                            # For subsequent HSN entries, don't repeat the invoice details
                            row = ["", "", "", "", "", hsndata.taxableamount, hsndata.tax_hsnsac_name,
                                   hsndata.cgst if taxtype == 'GST' else '',
                                   hsndata.cgst_amount if taxtype == 'GST' else '',
                                   hsndata.sgst if taxtype == 'GST' else '',
                                   hsndata.sgst_amount if taxtype == 'GST' else '',
                                   hsndata.sgst*2 if taxtype == 'IGST' else '',
                                   hsndata.cgst_amount*2 if taxtype == 'IGST' else '']

                        # Write the row to the sheet
                        for col_idx, cell_value in enumerate(row):
                            sheet.write(row_idx, col_idx, cell_value)

                            # Update max_col_widths for the current column if necessary
                            max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

                        row_idx += 1  # Move to the next row for the next entry

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 6000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{startdate} TO {enddate}Purches_invoices.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)





     