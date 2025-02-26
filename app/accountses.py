from django.shortcuts import render, redirect,HttpResponse 
from . models import *
from django.contrib import messages
from django.utils import timezone
from datetime import datetime
import datetime

from datetime import timedelta
from django.utils import timezone
from django.db.models import Sum
from django.db.models import F
from calendar import monthrange,month_name

import datetime
from datetime import  timedelta
from django.db.models import Q
from datetime import datetime

def accounts(request):
    try:
        if request.user.is_authenticated :
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  

            today = datetime.now().date()
            yestarday = today - timedelta(days=30)

            startdate = yestarday
            enddate = today 

            taxes = taxSlab.objects.filter(vendor=user,invoice__invoice_status=True,invoice__invoice_date__range=[startdate,enddate]).exclude(tax_rate_name='GST0').values('tax_rate_name', 'cgst').annotate(
                total_amount=Sum('total_amount')
            )
            
            
            Type="Sales"
            print(enddate)
            return render(request,'account.html',{'active_page': 'accounts','taxes':taxes,
                        'startdate':startdate,'enddate':enddate,'Type':Type})

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



def searchtaxesaccount(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')
            Type = request.POST.get('Type')
            print(Type,enddate,startdate)
            if Type=="Sales":
                taxes = taxSlab.objects.filter(vendor=user,invoice__invoice_status=True,invoice__invoice_date__range=[startdate,enddate]).exclude(tax_rate_name='GST0').values('tax_rate_name', 'cgst').annotate(
                    total_amount=Sum('total_amount')
                )
            else:
                 print("purchase")
                 taxes=None

            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            return render(request,'account.html',{'active_page': 'accounts','taxes':taxes,
                        'startdate':startdate,'enddate':enddate,'Type':Type})

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)


def searchtaxslabvidget(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')
            Type = request.POST.get('type')
            taxrate = float(request.POST.get('taxrate'))
            print(Type,enddate,startdate,taxrate)
            if Type=="Sales":
                taxes = taxSlab.objects.filter(vendor=user,cgst=taxrate).last()
                invoiceitem = InvoiceItem.objects.filter(vendor=user,invoice__invoice_date__range=[startdate,enddate],
                    invoice__invoice_status=True,cgst_rate=taxrate)
                
                print(invoiceitem)
            else:
                 print("purchase")
                 taxes=None

         
            return render(request,'taxdetailsitem.html',{'active_page': 'accounts','taxes':taxes,
                        'startdate':startdate,'enddate':enddate,'Type':Type,'invoiceitem':invoiceitem})
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)


def gstr1(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')

            b2bCount = Invoice.objects.filter(vendor=user,invoice_status=True,invoice_date__range=[startdate,enddate]).exclude(customer_gst_number="").count()

            Invoiceamount = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate]
                            ).exclude(customer_gst_number=""
                            ).aggregate(total=Sum('grand_total_amount'))['total'] or 0

            taxable_total_amount = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate]
                            ).exclude(customer_gst_number=""
                            ).aggregate(total=Sum('taxable_amount'))['total'] or 0
            
            CGST_total_amount = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate]
                            ).exclude(customer_gst_number=""
                            ).aggregate(total=Sum('sgst_amount'))['total'] or 0

            if CGST_total_amount:
                totaltaxamount = CGST_total_amount * 2
            else:
                totaltaxamount = 0

            # b2c work start here

            b2cCount = Invoice.objects.filter(vendor=user,invoice_status=True,invoice_date__range=[startdate,enddate],customer_gst_number="").count()

            b2cInvoiceamount = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate],customer_gst_number=""
                            ).aggregate(total=Sum('grand_total_amount'))['total'] or 0
            
            b2ctaxable_total_amount = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate],customer_gst_number=""
                            ).aggregate(total=Sum('taxable_amount'))['total'] or 0
            
            b2cCGST_total_amount = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate],customer_gst_number=""
                            ).aggregate(total=Sum('sgst_amount'))['total'] or 0
            
            if b2cCGST_total_amount:
                b2ctotaltaxamount = b2cCGST_total_amount * 2
            else:
                b2ctotaltaxamount = 0


            # final total calculation

            invoicefilterd = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate])  

            nillrateamount = InvoiceItem.objects.filter(cgst_rate__lt=1,invoice__in=invoicefilterd
                                ).aggregate(total=Sum('total_amount'))['total'] or 0        

            totalvouchercount = b2bCount + b2cCount
            totaltaxableamount = taxable_total_amount + b2ctaxable_total_amount + nillrateamount
            cgsttotalboth = CGST_total_amount + b2cCGST_total_amount
            totaltaxamountboth = totaltaxamount + b2ctotaltaxamount
            totalinvoiceamounts = Invoiceamount + b2cInvoiceamount

            # nill rated nvoice data 


            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            return render(request,'gstr1.html',{'active_page': 'accounts',
                        'startdate':startdate,'enddate':enddate,'b2bCount':b2bCount,
                        'taxable_total_amount':taxable_total_amount,'CGST_total_amount':CGST_total_amount,
                        'totaltaxamount':totaltaxamount,'Invoiceamount':Invoiceamount,'b2cCount':b2cCount,
                        'b2cInvoiceamount':b2cInvoiceamount,'b2ctaxable_total_amount':b2ctaxable_total_amount,
                        'b2cCGST_total_amount':b2cCGST_total_amount,'b2ctotaltaxamount':b2ctotaltaxamount,
                        'totalvouchercount':totalvouchercount,'totaltaxableamount':totaltaxableamount,
                        'cgsttotalboth':cgsttotalboth,'totaltaxamountboth':totaltaxamountboth,
                        'totalinvoiceamounts':totalinvoiceamounts,'nillrateamount':nillrateamount})
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)




def b2bInvoicedetails(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')

            b2binvoice = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate]
                            ).exclude(customer_gst_number="")

            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            return render(request,'b2binvc.html',{'active_page': 'accounts',
                        'startdate':startdate,'enddate':enddate,'b2binvoice':b2binvoice})
        
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)

        
def nillratedforms(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')

            invoicefilterd = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate]) 

            nillrateamount = InvoiceItem.objects.filter(cgst_rate__lt=1,invoice__in=invoicefilterd)

            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            return render(request,'nillrated.html',{'active_page': 'accounts',
                        'startdate':startdate,'enddate':enddate,'nillrateamount':nillrateamount})
        
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)


def b2cInvoicetaxdetails(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')

            b2cinvoice = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate],customer_gst_number=""
                            )
            
            # tax 12
            tax12taxableamount = InvoiceItem.objects.filter(invoice__in=b2cinvoice,cgst_rate=6.00
                        ).aggregate(total=Sum('totalwithouttax'))['total'] or 0 
            
            tax12cgstamount = InvoiceItem.objects.filter(invoice__in=b2cinvoice,cgst_rate=6.00
                        ).aggregate(total=Sum('cgst_rate_amount'))['total'] or 0 
            
            tax12totalamount = tax12cgstamount *2 + tax12taxableamount

            # tax 18

            tax18taxableamount = InvoiceItem.objects.filter(invoice__in=b2cinvoice,cgst_rate=9.00
                        ).aggregate(total=Sum('totalwithouttax'))['total'] or 0 
            
            tax18cgstamount = InvoiceItem.objects.filter(invoice__in=b2cinvoice,cgst_rate=9.00
                        ).aggregate(total=Sum('cgst_rate_amount'))['total'] or 0 
            
            tax18totalamount = tax18cgstamount *2 + tax18taxableamount

            # tax 5

            tax5taxableamount = InvoiceItem.objects.filter(invoice__in=b2cinvoice,cgst_rate=2.50
                        ).aggregate(total=Sum('totalwithouttax'))['total'] or 0 
            
            tax5cgstamount = InvoiceItem.objects.filter(invoice__in=b2cinvoice,cgst_rate=2.50
                        ).aggregate(total=Sum('cgst_rate_amount'))['total'] or 0 
            
            tax5totalamount = tax5cgstamount *2 + tax5taxableamount


            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            return render(request,'b2cdetails.html',{'active_page': 'accounts',
                        'startdate':startdate,'enddate':enddate,'tax12taxableamount':tax12taxableamount,
                        'tax12cgstamount':tax12cgstamount,'tax12totalamount':tax12totalamount,
                        'tax18taxableamount':tax18taxableamount,'tax18cgstamount':tax18cgstamount,
                        'tax18totalamount':tax18totalamount,'tax5taxableamount':tax5taxableamount,
                        'tax5cgstamount':tax5cgstamount,'tax5totalamount':tax5totalamount})
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
# gstr1 excet donwload code not in live uses
def generate_gstr1_excel(request):
    # Create an Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Sample data for multiple sheets
    data = {
        "b2b,sez,de": [["GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice date",'Invoice Value','Place Of Supply','Reverse Charge','Applicable % of Tax Rate','Invoice Type','E-commerce GSTIN','Rate','Taxable Value','Cess Amount']],
        "b2ba": [["Place Of Supply", "Reverse Charge", "Applicable % of Tax Rate", "Invoice Type","E-Commerce GSTIN","Rate",'Taxable Value','Cess Amount']],
        "b2cl": [["Invoice Number", "Invoice date", "Invoice Value","Place Of Supply","Applicable % of Tax Rate",'Rate','Taxable Value','Cess Amount',"E-Commerce GSTIN"]],
        "b2cla": [["Original Invoice Number", "Original Invoice date", "Original Place Of Supply", "Revised Invoice Number", "Revised Invoice date","Invoice Value","Applicable % of Tax Rate","Rate",'Taxable Value','Cess Amount',"E-Commerce GSTIN"]],
        "b2cs": [["Type","Place Of Supply","Applicable % of Tax Rate","Rate",'Taxable Value','Cess Amount','E-Commerce GSTIN']],
        "b2csa": [["Financial Year","Original Month","Place Of Supply","Type","Applicable % of Tax Rate","Rate",'Taxable Value','Cess Amount','E-Commerce GSTIN']],
        "cdnr": [["GSTIN/UIN of Recipient", "Receiver Name", "Note Number", "Note Date",'Note Type','Place Of Supply','Reverse Charge','Note Supply Type','Note Value','Applicable % of Tax Rate','Rate','Taxable Value','Cess Amount']],
        "cdnra": [["GSTIN/UIN of Recipient", "Receiver Name", "Original Note Number", "Original Note Date","Revised Note Number","Revised Note Date",'Note Type','Place Of Supply','Reverse Charge','Note Supply Type','Note Value','Applicable % of Tax Rate','Rate','Taxable Value','Cess Amount']],
        "cdnur": [["UR Type","Note Number","Note Date",'Note Type','Note Value',"Applicable % of Tax Rate","Rate",'Taxable Value','Cess Amount']],
        "cdnura": [["UR Type","Original Note Number","Original Note Date","Revised Note Number","Revised Note Date",'Note Type','Place Of Supply','Note Value',"Applicable % of Tax Rate","Rate",'Taxable Value','Cess Amount']],
        "exp": [['Export Type',"Invoice Number", "Invoice date", "Invoice Value","Port Code","Shipping Bill Number","Shipping Bill Date",'Rate','Taxable Value','Cess Amount']],
        "expa": [["Revised Invoice date", "Invoice Value","Port Code","Shipping Bill Number","Shipping Bill Date",'Rate','Taxable Value','Cess Amount']],
        "at": [['Place Of Supply', "Applicable % of Tax Rate","Rate","Gross Advance Received",'Cess Amount']],
        "ata": [["Financial Year","Original Month","Original Place Of Supply","Applicable % of Tax Rate",'Rate',"Gross Advance Received",'Cess Amount']],
        "atadj": [['Place Of Supply', "Applicable % of Tax Rate","Rate","Gross Advance Adjusted",'Cess Amount']],
        "atadja": [["Financial Year","Original Month","Original Place Of Supply","Applicable % of Tax Rate","Rate","Gross Advance Adjusted",'Cess Amount']],
        "exemp": [['Description', "Nil Rated Supplies","Exempted(other than nil Rated/non GST supply)",'Non-GST supplies ']],
        "hsn": [['HSN', "Description","UQC","Total Quantity",'Total Value','Rate','Taxable Value','Integrated Tax Amount','Central Tax Amount','State/UT Tax Amount','Cess Amount']],
        "docs": [['Nature of Document', "Sr. No. Form","Sr. No. To",'Total Number','Total Cancelled']],
    }

    # Adding data to each sheet
    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name)
        
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}{row_idx}"] = value

    # Create HTTP response with Excel content
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="GSTR1_Report.xlsx"'

    # Save workbook to response
    wb.save(response)
    return response



def onlinechannel(request):
    try:
        if request.user.is_authenticated :
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            onlinechannel =onlinechannls.objects.filter(vendor=user)
            return render(request,'onlinechannel.html',{'onlinechannel':onlinechannel})
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)

        

def addchannel(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            chanelname = request.POST.get('chanelname')
            channelgstin = request.POST.get('channelgstin')
            if onlinechannls.objects.filter(vendor=user,channalname=chanelname).exists():
                messages.error(request,'channel already exists!')

            else:
                onlinechannls.objects.create(vendor=user,channalname=chanelname,company_gstin=channelgstin)
                messages.success(request,'channel Created!')

            return redirect('onlinechannel')
        
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)

        

def deletechannel(request,id):
    try:
        if request.user.is_authenticated:
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            id=id
            if onlinechannls.objects.filter(vendor=user,id=id).exists():
                onlinechannls.objects.filter(vendor=user,id=id).delete()
                messages.success(request,'channel Deleted!')

            else:
                messages.success(request,'Channel Not exists!')

            return redirect('onlinechannel')
        
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



def updatechanel(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            mainids = request.POST.get('mainids')
            chanelname = request.POST.get('chanelname')
            channelgstin = request.POST.get('channelgstin')
            if onlinechannls.objects.filter(id=mainids).exists():
                onlinechannls.objects.filter(id=mainids).update(channalname=chanelname,company_gstin=channelgstin)
                messages.success(request,'channel Updated!')

            else:
                onlinechannls.objects.create(vendor=user,channalname=chanelname,company_gstin=channelgstin)
                messages.error(request,'channel Not exists!')

            return redirect('onlinechannel')
        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



def proformainvoice(request,id):
    try:
        if request.user.is_authenticated:
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            userid = id
            guestdata = Gueststay.objects.filter(vendor=user, id=userid)
            invoice_data = Invoice.objects.get(vendor=user, customer=userid)
            profiledata = HotelProfile.objects.filter(vendor=user)
            itemid = invoice_data.id
            status = invoice_data.foliostatus

            invoice_datas = Invoice.objects.filter(vendor=user, customer=userid)
            invoiceitemdata = InvoiceItem.objects.filter(vendor=user, invoice=itemid).order_by('id')
            loyltydata = loylty_data.objects.filter(vendor=user, Is_active=True)
            invcpayments = InvoicesPayment.objects.filter(vendor=user,invoice=itemid).all()
            taxelab = taxSlab.objects.filter(vendor=user,invoice=itemid)
           
 
            if invoice_data.taxtype == 'GST':
                    gstamounts = invoice_data.gst_amount
                    sstamounts = invoice_data.sgst_amount
                    checkproforma= "Pro Forma Invoice"
                    invcheck =  invoiceDesign.objects.get(vendor=user)
                    if invcheck.guestinvcdesign==1:
                        return render(request, 'invoicepage.html', {
                            'profiledata': profiledata,
                            'guestdata': guestdata,
                            'invoice_data': invoice_datas,
                            'invoiceitemdata': invoiceitemdata,
                            'invcpayments':invcpayments,
                            'gstamounts':gstamounts,
                            'sstamounts':sstamounts,
                            'taxelab':taxelab,
                            'checkproforma':checkproforma
                        })
                    elif invcheck.guestinvcdesign==2:
                        return render(request, 'invoicepage2.html', {
                            'profiledata': profiledata,
                            'guestdata': guestdata,
                            'invoice_data': invoice_datas,
                            'invoiceitemdata': invoiceitemdata,
                            'invcpayments':invcpayments,
                            'gstamounts':gstamounts,
                            'sstamounts':sstamounts,
                            'taxelab':taxelab,
                            'checkproforma':checkproforma
                        })
                        
              
        else:
            return render(request, 'login.html')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



def hsnsummry(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')

            invoices = Invoice.objects.filter(vendor=user,invoice_status=True,
                            invoice_date__range=[startdate,enddate]
                            )



            hsn_data = InvoiceItem.objects.filter(invoice__in=invoices).values('hsncode', 'cgst_rate', 'sgst_rate').annotate(
                total_amount=Sum('total_amount'),
                totalwithouttax=Sum('totalwithouttax'),
                total_tax_rate=F('cgst_rate') + F('sgst_rate')  # Each row keeps its own tax rate
            )

            
            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            
            return render(request,'hsnsummry.html',{'hsn_data':hsn_data,'startdate':startdate,'enddate':enddate})

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



from django.db.models import Min, Max
def documentsummry(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor  
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')



            invoices = Invoice.objects.filter(
                vendor=user,
                invoice_status=True,
                invoice_date__range=[startdate, enddate]
            )

            totalinvoices = invoices.count()

            first_invoice_number = invoices.aggregate(
                min_invoice=Min('invoice_number')
            )['min_invoice']

            last_invoice_number = invoices.aggregate(
                max_invoice=Max('invoice_number')
            )['max_invoice']

            print(f"Total Invoices: {totalinvoices}")
            print(f"First Invoice Number on {startdate}: {first_invoice_number}")
            print(f"Last Invoice Number on {enddate}: {last_invoice_number}")



            hotelnamedata = HotelProfile.objects.get(vendor=user)
            hotelname = hotelnamedata.name
            hoteltype = "hotel rooms"


            startdate = datetime.strptime(startdate, "%Y-%m-%d")
            enddate = datetime.strptime(enddate, "%Y-%m-%d")
            
            return render(request,'documentsummry.html',{'startdate':startdate,'enddate':enddate,'totalinvoices':totalinvoices,
                            'hotelname':hotelname,'hoteltype':hoteltype,'firstinvoicenumer':first_invoice_number,'last_invoice_number':last_invoice_number})

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



import calendar
from django.contrib.auth.decorators import login_required
import xlwt

# gstr-1 mix excel donwload
@login_required
def generate_gstr1_mix_invoice_excel(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')
           

            # Filter invoices based on the given month and invoice_status
            invoices = Invoice.objects.filter(
                invoice_date__range=[startdate, enddate],
                invoice_status=True,
                vendor=user
            ).order_by('invoice_number')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet("GSTR-1 Invoices")

            # Define the header
            # headers = ["Invoice Number", "Guest Name", "Phone", "Source" ,"Check-in Date", "Check-out Date",
            #         "Tax Amount", "Grand Total Amount",  "Tax Type", "GST Number"]

            headers = ["Type","Hotel State","GSTIN/NUM","Guest Name","Check-in Date","Check-out Date","Source","Guest GST",
                       "Guest Company","Billed Date","Invoice Number","Taxable Amount","Integrated Tax Amount",
                       "Central Tax Amount","State Tax Amount","Cess  Amount","Tax Amount","Invoice  Amount"]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]

            hoteldata = HotelProfile.objects.get(vendor=user)
            # Write the invoice data
            for row_idx, invoice in enumerate(invoices, start=1):
                guest = invoice.customer
                if invoice.customer_gst_number=='':
                    types="B2C"
                elif invoice.customer_gst_number is not None:
                    types="B2B"
                row = [
                    types,  # Convert invoice_number to string to handle longer values
                    hoteldata.zipcode,
                    hoteldata.gstin,  # Convert guestphome to string
                    guest.guestname,
                    guest.checkindate.strftime('%Y-%m-%d %H:%M'),  # Format with date and time
                    guest.checkoutdate.strftime('%Y-%m-%d %H:%M'),  # Format with date and time
                    guest.channel,
                    invoice.customer_gst_number,
                    invoice.customer_company,
                    invoice.invoice_date.strftime('%Y-%m-%d'),
                    invoice.invoice_number,
                    invoice.taxable_amount,
                    0,
                    invoice.gst_amount,
                    invoice.sgst_amount,
                    '',
                    float(invoice.gst_amount * 2),  # Example calculation, adjust as needed
                    float(invoice.grand_total_amount),
                ]

                for col_idx, cell_value in enumerate(row):
                    sheet.write(row_idx, col_idx, cell_value)

                    # Update max_col_widths for the current column if necessary
                    max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 8000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="Invoices_export{startdate} TO {enddate}.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



def generate_b2b_invoice_excel(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            startdate = request.POST.get('startdate')
            enddate = request.POST.get('enddate')
           

            # Filter invoices based on the given month and invoice_status
            invoices = Invoice.objects.filter(
                invoice_date__range=[startdate, enddate],
                invoice_status=True,
                vendor=user
            ).exclude(customer_gst_number="").order_by('invoice_number')

            # Create a new Workbook
            workbook = xlwt.Workbook(encoding='utf-8')
            sheet = workbook.add_sheet("GSTR-1 Invoices")

            # Define the header
            # headers = ["Invoice Number", "Guest Name", "Phone", "Source" ,"Check-in Date", "Check-out Date",
            #         "Tax Amount", "Grand Total Amount",  "Tax Type", "GST Number"]

            headers = ["Guest Name","Check-in Date","Check-out Date","Source","Particular","GSTIN/NUM",
                       "Taxable Amount","Integrated Tax Amount",
                       "Central Tax Amount","State Tax Amount","Cess  Amount","Tax Amount","Invoice  Amount"]

            # Write the header row
            for col_idx, header in enumerate(headers):
                sheet.write(0, col_idx, header)

            # Initialize list to track maximum content lengths for each column
            max_col_widths = [len(header) for header in headers]

            hoteldata = HotelProfile.objects.get(vendor=user)
            # Write the invoice data
            for row_idx, invoice in enumerate(invoices, start=1):
                guest = invoice.customer
                row = [
                    guest.guestname,  # Convert invoice_number to string to handle longer values
                    guest.checkindate.strftime('%Y-%m-%d %H:%M'),  # Format with date and time
                    guest.checkoutdate.strftime('%Y-%m-%d %H:%M'),  # Format with date and time
                    guest.channel,
                    invoice.customer_company,
                    invoice.customer_gst_number,
                    invoice.taxable_amount,
                    0,
                    invoice.gst_amount,
                    invoice.sgst_amount,
                    '',
                    float(invoice.gst_amount * 2),  # Example calculation, adjust as needed
                    float(invoice.grand_total_amount),
                ]

                for col_idx, cell_value in enumerate(row):
                    sheet.write(row_idx, col_idx, cell_value)

                    # Update max_col_widths for the current column if necessary
                    max_col_widths[col_idx] = max(max_col_widths[col_idx], len(str(cell_value)))

            # Auto-adjust column widths based on content size
            for col_idx, max_length in enumerate(max_col_widths):
                sheet.col(col_idx).width = max(max_length * 256, 6000)  # Minimum width of 8000

            # Create HttpResponse with content type for Excel .xls file
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="B2B Details {startdate} TO {enddate}_invoices.xls"'

            # Save workbook to response
            workbook.save(response)
            return response

        else:
            return redirect('loginpage')
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)



def createpurchasefromproduct(request):
    try:
        if request.user.is_authenticated and request.method == "POST":
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 
            pid = request.POST.get('pid')
            add_qty = request.POST.get('add_qty')
            totalamount = request.POST.get('totalamount')
            print(pid,add_qty,totalamount)

            itempurchasehistorlocal.objects.create(vendor=user,items_id=pid,add_qty=add_qty,total_amount=totalamount)
            
            Items.objects.filter(vendor=user,id=pid).update(
                available_qty=F('available_qty')+int(add_qty),
                total_qty=F('total_qty')+int(add_qty)
            )
            
            messages.success(request,"Succefully Created Purchase For this day!")

            return redirect('Product')
        else:
            return redirect('loginpage')
        
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)


def payble(request):
    try:
        if request.user.is_authenticated :
            user = request.user
            subuser = Subuser.objects.select_related('vendor').filter(user=user).first()
            if subuser:
                user = subuser.vendor 

            purchasepanding = Supplier.objects.filter(vendor=user,unpaid=True)

            totaldueamount = Supplier.objects.filter(vendor=user,unpaid=True
                        ).aggregate(total=Sum('due_amount'))['total'] or 0 

            return render(request,'payables.html',{'purchasepanding':purchasepanding,'totaldueamount':totaldueamount})
        
        else:
            return redirect('loginpage')
        
    except Exception as e:
        return render(request, '404.html', {'error_message': str(e)}, status=500)

     
